import json
from pathlib import Path

import openpyxl
import pyarrow.parquet as pq

from geologparser.export import tabular_rows, write_geoparquet, write_parquet_dataset, write_xlsx


ROOT = Path(__file__).resolve().parents[1]


def record():
    return json.loads((ROOT / "examples/boreholes/synthetic_valid.json").read_text(encoding="utf-8"))


def test_tabular_rows_keep_raw_normalized_and_provenance():
    source = record()
    rows = tabular_rows([source])
    assert rows["boreholes"][0]["document_id"] == source["document"]["document_id"]
    assert rows["intervals"][0]["lithology_raw"] == source["intervals"][0]["lithology_raw"]["value"]
    provenance = {row["field_path"]: row for row in rows["provenance"]}
    assert provenance["intervals[0].bottom_depth_m"]["source_text"] == source["intervals"][0]["bottom_depth_m"]["source_text"]


def test_tabular_rows_preserve_human_display_bbox_provenance():
    source = record()
    source["borehole"]["final_depth_m"].update({
        "display_bbox": [1, 2, 30, 40], "display_bbox_source": "human_drawn",
        "display_bbox_annotator_id": "reviewer-1",
    })
    rows = tabular_rows([source])
    row = next(item for item in rows["provenance"] if item["field_path"] == "borehole.final_depth_m")
    assert row["display_bbox_source"] == "human_drawn"
    assert row["display_bbox_annotator_id"] == "reviewer-1"


def test_xlsx_has_three_traceable_sheets(tmp_path):
    path = tmp_path / "boreholes.xlsx"
    write_xlsx([record()], path)
    workbook = openpyxl.load_workbook(path, read_only=True)
    assert workbook.sheetnames == ["boreholes", "intervals", "provenance"]
    assert workbook["intervals"].max_row == len(record()["intervals"]) + 1


def test_parquet_dataset_preserves_tables_and_empty_intervals(tmp_path):
    source = record()
    source["intervals"] = []
    directory = tmp_path / "parquet"
    write_parquet_dataset([source], directory)
    assert pq.read_table(directory / "boreholes.parquet").num_rows == 1
    interval_table = pq.read_table(directory / "intervals.parquet")
    assert interval_table.num_rows == 0
    assert "interval_id" in interval_table.column_names
    assert json.loads((directory / "metadata.json").read_text(encoding="utf-8"))["format"].endswith("v001")


def test_geoparquet_has_geo_metadata_and_rejects_mixed_crs(tmp_path):
    first = record()
    first["borehole"]["x_coordinate"]["value"] = 329168
    first["borehole"]["y_coordinate"]["value"] = 405889
    first["borehole"]["coordinate_system"]["value"] = "EPSG:27700"
    path = tmp_path / "boreholes.parquet"
    write_geoparquet([first], path)
    metadata = pq.read_schema(path).metadata
    geo = json.loads(metadata[b"geo"])
    assert geo["primary_column"] == "geometry"
    assert geo["columns"]["geometry"]["crs"]["id"] == {"authority": "EPSG", "code": 27700}
    second = record()
    second["document"]["document_id"] = "other"
    second["borehole"]["x_coordinate"]["value"] = 1
    second["borehole"]["y_coordinate"]["value"] = 2
    second["borehole"]["coordinate_system"]["value"] = "EPSG:4326"
    import pytest
    with pytest.raises(ValueError, match="exactly one"):
        write_geoparquet([first, second], tmp_path / "mixed.parquet")
