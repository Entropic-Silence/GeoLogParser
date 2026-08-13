from pathlib import Path

from openpyxl import Workbook
import pytest

from geologparser.evaluation import (
    SurfacePoint,
    convex_hull_xy,
    load_coal_602_roof_depth_surface,
    perturb_surface_scalar,
    regular_queries_within_hull,
)


def _workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "Gas Drainage Borehole Number", "Y", "X", "Z", "Drilling Registration",
        None, None, "Drilling Measurement", None, "Drilling, coal seam parameters",
        None, None, None,
    ])
    sheet.append([
        None, None, None, None, "Final hole depth(m)", "3 Coal Roof Depth(m)",
        "3 Coal seam thickness(m)", "Borehole Zenith Angle (°)",
        "Drilling azimuth (°)", "borehole inclina-tion (°)",
        "borehole azimuth (°)", "coal seam dip angle (°)",
        "coal seam dip direction (°)",
    ])
    sheet.append(["PRIVATE-A", 100, 200, 30, 50, 40, 3, 120, 90, 30, 90, 4, 3.5])
    sheet.append(["PRIVATE-B", 110, 200, 31, 51, 41, 3, 120, 90, 30, 90, 4, 3.5])
    sheet.append(["PRIVATE-C", 100, 220, 32, 52, 42, 3, 120, 90, 30, 90, 4, 3.5])
    workbook.save(path)


def test_load_source_surface_suppresses_origin_and_identifiers(tmp_path: Path):
    path = tmp_path / "coal.xlsx"
    _workbook(path)
    surface = load_coal_602_roof_depth_surface(path)
    assert surface.source_record_count == 3
    assert surface.coordinate_extent_u_m == 10
    assert surface.coordinate_extent_v_m == 20
    assert surface.scalar_minimum_m == 40
    assert surface.scalar_maximum_m == 42
    assert surface.coordinate_origin_persisted is False
    assert {(point.x, point.y) for point in surface.points} == {(0, 0), (10, 0), (0, 20)}
    assert [point.borehole_id for point in surface.points] == [
        "SOURCE_ROW_0001", "SOURCE_ROW_0002", "SOURCE_ROW_0003",
    ]
    assert all("PRIVATE" not in point.borehole_id for point in surface.points)


def test_hull_grid_excludes_bounding_box_points_outside_triangle():
    points = (
        SurfacePoint(0, 0, 1, "a"),
        SurfacePoint(10, 0, 1, "b"),
        SurfacePoint(0, 10, 1, "c"),
        SurfacePoint(2, 2, 1, "d"),
    )
    hull = convex_hull_xy(points)
    assert hull == ((0.0, 0.0), (10.0, 0.0), (0.0, 10.0))
    queries = regular_queries_within_hull(hull, 3)
    assert (10.0, 10.0) not in queries
    assert {(0.0, 0.0), (5.0, 5.0), (10.0, 0.0)} <= set(queries)


def test_source_scalar_perturbation_is_seeded_and_preserves_geometry():
    points = (
        SurfacePoint(0, 0, 10, "a"), SurfacePoint(1, 0, 20, "b"),
    )
    first = perturb_surface_scalar(points, 0.5, 7)
    second = perturb_surface_scalar(points, 0.5, 7)
    assert first == second
    assert [(point.x, point.y, point.borehole_id) for point in first] == [
        (0, 0, "a"), (1, 0, "b"),
    ]
    assert all(abs(after.elevation - before.elevation) == 0.5 for before, after in zip(points, first))
    with pytest.raises(ValueError, match="non-negative"):
        perturb_surface_scalar(points, -0.1, 7)


def test_source_surface_rejects_duplicate_coordinates(tmp_path: Path):
    path = tmp_path / "coal.xlsx"
    _workbook(path)
    from openpyxl import load_workbook
    workbook = load_workbook(path)
    sheet = workbook.active
    sheet["B4"] = 100
    sheet["C4"] = 200
    workbook.save(path)
    with pytest.raises(ValueError, match="duplicate collar coordinates"):
        load_coal_602_roof_depth_surface(path)
