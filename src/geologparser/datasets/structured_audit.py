"""Immutable, hash-bound audits for acquired structured borehole datasets."""

from __future__ import annotations

from collections import Counter
from datetime import datetime, timezone
import hashlib
import json
import math
import os
from pathlib import Path
import re
import shutil
import tempfile
from typing import Any, Iterable


AUDIT_SCHEMA_VERSION = "structured_borehole_content_audit_v001"
SUPPORTED_PROFILES = {"binhai_cptu_v001", "coal_boreholes_602_v001"}


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _json_bytes(value: Any) -> bytes:
    return (json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n").encode("utf-8")


def _finite_number(value: Any) -> float | None:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return None
    number = float(value)
    return number if math.isfinite(number) else None


def _range(values: Iterable[float]) -> dict[str, float | int | None]:
    rows = list(values)
    return {
        "count": len(rows),
        "minimum": min(rows) if rows else None,
        "maximum": max(rows) if rows else None,
    }


def _load_acquisition(dataset_root: Path) -> tuple[dict[str, Any], dict[str, dict[str, Any]]]:
    path = dataset_root / "metadata/acquisition.json"
    evidence = json.loads(path.read_text(encoding="utf-8"))
    rows = evidence.get("files")
    if not isinstance(rows, list) or not rows:
        raise ValueError("acquisition evidence has no files")
    inventory: dict[str, dict[str, Any]] = {}
    for row in rows:
        filename = str(row.get("filename") or "")
        if not filename or Path(filename).name != filename or filename in inventory:
            raise ValueError("acquisition evidence contains an unsafe or duplicate filename")
        source = dataset_root / "raw" / filename
        if not source.is_file():
            raise ValueError(f"acquired source file is missing: {filename}")
        if source.stat().st_size != row.get("size_bytes") or _sha256_file(source) != row.get("sha256"):
            raise ValueError(f"acquired source file does not match evidence: {filename}")
        inventory[filename] = row
    actual_files = {path.name for path in (dataset_root / "raw").iterdir() if path.is_file()}
    if actual_files != set(inventory):
        raise ValueError("acquired raw-file set does not match acquisition evidence")
    return evidence, inventory


def _scan_text_privacy_signals(dataset_root: Path, inventory: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    patterns = {
        "email_address": re.compile(r"\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b", re.IGNORECASE),
        "telephone_like_number": re.compile(r"(?<!\d)(?:\+?86[- ]?)?1[3-9]\d{9}(?!\d)"),
    }
    findings: list[dict[str, Any]] = []
    for filename, row in sorted(inventory.items()):
        content_type = str(row.get("content_type") or "")
        if not (content_type.startswith("text/") or Path(filename).suffix.lower() in {".md", ".txt", ".py"}):
            continue
        text = (dataset_root / "raw" / filename).read_text(encoding="utf-8", errors="replace")
        for category, pattern in patterns.items():
            count = len(pattern.findall(text))
            if count:
                findings.append({"category": category, "file": filename, "match_count": count})
    return findings


def _openpyxl():
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - environment-dependent guard
        raise RuntimeError("structured workbook auditing requires openpyxl") from exc
    return load_workbook


def _audit_binhai(dataset_root: Path, inventory: dict[str, dict[str, Any]]) -> dict[str, Any]:
    load_workbook = _openpyxl()
    expected_header = ("Name", "X", "Y", "Depth", "qt", "fs", "u2")
    workbook_names = sorted(name for name in inventory if Path(name).suffix.lower() == ".xlsx")
    unexpected_files = sorted(name for name in inventory if name not in workbook_names)
    files: list[dict[str, Any]] = []
    global_missing: Counter[str] = Counter()
    global_ranges: dict[str, list[float]] = {name: [] for name in ("Depth", "qt", "fs", "u2")}
    total_rows = 0
    identity_mismatches = 0
    non_increasing_depth_steps = 0
    coordinate_values: dict[str, set[str]] = {"X": set(), "Y": set()}

    for filename in workbook_names:
        path = dataset_root / "raw" / filename
        workbook = load_workbook(path, read_only=True, data_only=False)
        if len(workbook.sheetnames) != 1:
            raise ValueError(f"Binhai workbook must contain one sheet: {filename}")
        sheet = workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        header = tuple(next(rows, ()))
        expected_name = Path(filename).stem.removesuffix("_update")
        file_missing: Counter[str] = Counter()
        file_ranges: dict[str, list[float]] = {name: [] for name in ("Depth", "qt", "fs", "u2")}
        file_rows = 0
        file_identity_mismatches = 0
        file_non_increasing = 0
        previous_depth: float | None = None
        depth_steps: Counter[str] = Counter()

        for row in rows:
            file_rows += 1
            padded = tuple(row) + (None,) * max(0, len(expected_header) - len(row))
            values = dict(zip(expected_header, padded))
            if values["Name"] != expected_name:
                file_identity_mismatches += 1
            for field in expected_header:
                if values[field] is None:
                    file_missing[field] += 1
            for field in ("Depth", "qt", "fs", "u2"):
                number = _finite_number(values[field])
                if number is not None:
                    file_ranges[field].append(number)
                    global_ranges[field].append(number)
            for field in ("X", "Y"):
                if values[field] is not None:
                    coordinate_values[field].add(str(values[field]))
            depth = _finite_number(values["Depth"])
            if depth is not None and previous_depth is not None:
                step = depth - previous_depth
                depth_steps[f"{step:.12g}"] += 1
                if step <= 0:
                    file_non_increasing += 1
            if depth is not None:
                previous_depth = depth

        workbook.close()
        total_rows += file_rows
        identity_mismatches += file_identity_mismatches
        non_increasing_depth_steps += file_non_increasing
        global_missing.update(file_missing)
        files.append({
            "filename": filename,
            "sha256": inventory[filename]["sha256"],
            "sheet_name": sheet.title,
            "worksheet_dimensions": {"rows_including_header": sheet.max_row, "columns": sheet.max_column},
            "header": list(header),
            "header_matches_profile": header == expected_header,
            "data_rows": file_rows,
            "missing_cells_by_field": dict(sorted(file_missing.items())),
            "numeric_ranges": {field: _range(values) for field, values in file_ranges.items()},
            "name_expected_from_filename": expected_name,
            "name_mismatch_rows": file_identity_mismatches,
            "non_increasing_depth_steps": file_non_increasing,
            "depth_step_counts": dict(sorted(depth_steps.items())),
        })

    coordinates_redacted = all(coordinate_values[field] == {"*"} for field in ("X", "Y"))
    all_headers_match = all(row["header_matches_profile"] for row in files)
    blockers = [
        "inventory_contains_cptu_workbooks_only_no_borehole_record_or_laboratory_files",
        "x_y_coordinates_redacted_as_asterisks",
        "coordinate_reference_system_not_declared_in_workbooks",
        "measurement_units_not_declared_in_workbooks",
        "no_lithology_or_stratigraphic_intervals",
        "not_ai_extraction_output_or_human_ground_truth",
    ]
    return {
        "profile": "binhai_cptu_v001",
        "content_classification": "structured_cptu_measurements",
        "source_data_role": "source_structured_data",
        "workbook_count": len(workbook_names),
        "unexpected_non_workbook_files": unexpected_files,
        "total_measurement_rows": total_rows,
        "schema": list(expected_header),
        "all_headers_match_profile": all_headers_match,
        "name_mismatch_rows": identity_mismatches,
        "non_increasing_depth_steps": non_increasing_depth_steps,
        "missing_cells_by_field": dict(sorted(global_missing.items())),
        "numeric_ranges": {field: _range(values) for field, values in global_ranges.items()},
        "coordinate_distinct_values": {field: sorted(values) for field, values in coordinate_values.items()},
        "coordinates_redacted": coordinates_redacted,
        "files": files,
        "eligibility": {
            "phase1_document_extraction": False,
            "paper1_benchmark": False,
            "paper2_method_evaluation": False,
            "paper3_spatial_error_propagation": False,
            "paper3_nonspatial_cptu_protocol_development": True,
        },
        "formal_use_status": "ineligible_for_current_formal_experiments",
        "formal_use_blockers": blockers,
    }


def _flatten_coal_header(sheet: Any) -> list[str]:
    first = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    second = list(next(sheet.iter_rows(min_row=2, max_row=2, values_only=True)))
    flattened = []
    for index in range(max(len(first), len(second))):
        top = first[index] if index < len(first) else None
        bottom = second[index] if index < len(second) else None
        flattened.append(str(bottom if bottom is not None else top or "").strip())
    return flattened


def _audit_coal(dataset_root: Path, inventory: dict[str, dict[str, Any]]) -> dict[str, Any]:
    load_workbook = _openpyxl()
    workbook_names = sorted(name for name in inventory if Path(name).suffix.lower() == ".xlsx")
    if len(workbook_names) != 1:
        raise ValueError("coal-602 profile requires exactly one XLSX workbook")
    filename = workbook_names[0]
    workbook = load_workbook(dataset_root / "raw" / filename, read_only=False, data_only=False)
    if len(workbook.sheetnames) != 1:
        raise ValueError("coal-602 workbook must contain exactly one sheet")
    sheet = workbook[workbook.sheetnames[0]]
    headers = _flatten_coal_header(sheet)
    expected_headers = [
        "Gas Drainage Borehole Number", "Y", "X", "Z", "Final hole depth(m)",
        "3 Coal Roof Depth(m)", "3 Coal seam thickness(m)", "Borehole Zenith Angle (°)",
        "Drilling azimuth (°)", "borehole inclina-tion (°)", "borehole azimuth (°)",
        "coal seam dip angle (°)", "coal seam dip direction (°)",
    ]
    rows = list(sheet.iter_rows(min_row=3, values_only=True))
    formula_cells = sum(
        1 for row in sheet.iter_rows() for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    )
    workbook.close()

    missing: Counter[str] = Counter()
    numeric_ranges: dict[str, list[float]] = {header: [] for header in expected_headers[1:]}
    identifiers: Counter[str] = Counter()
    roof_greater_than_final = 0
    roof_plus_thickness_greater_than_final = 0
    final_minus_roof_values: list[float] = []
    nonpositive_depth_or_thickness = 0
    angle_relation_max_error = 0.0
    azimuth_relation_max_error = 0.0
    for row in rows:
        padded = tuple(row) + (None,) * max(0, len(expected_headers) - len(row))
        identifiers[str(padded[0])] += 1
        for index, header in enumerate(expected_headers):
            if padded[index] is None:
                missing[header] += 1
            elif index:
                number = _finite_number(padded[index])
                if number is not None:
                    numeric_ranges[header].append(number)
        final_depth = _finite_number(padded[4])
        roof_depth = _finite_number(padded[5])
        seam_thickness = _finite_number(padded[6])
        if final_depth is not None and roof_depth is not None:
            roof_greater_than_final += int(roof_depth > final_depth)
            final_minus_roof_values.append(final_depth - roof_depth)
        if final_depth is not None and roof_depth is not None and seam_thickness is not None:
            roof_plus_thickness_greater_than_final += int(roof_depth + seam_thickness > final_depth)
            nonpositive_depth_or_thickness += int(
                final_depth <= 0 or roof_depth <= 0 or seam_thickness <= 0
            )
        zenith = _finite_number(padded[7])
        inclination = _finite_number(padded[9])
        if zenith is not None and inclination is not None:
            angle_relation_max_error = max(angle_relation_max_error, abs(zenith - inclination - 90.0))
        drilling_azimuth = _finite_number(padded[8])
        borehole_azimuth = _finite_number(padded[10])
        if drilling_azimuth is not None and borehole_azimuth is not None:
            azimuth_relation_max_error = max(azimuth_relation_max_error, abs(drilling_azimuth - borehole_azimuth))

    missing_script_references: list[str] = []
    run_all = dataset_root / "raw/run_all.py"
    if run_all.is_file():
        for relative in re.findall(r'root\s*/\s*"([^"]+)"\s*/\s*"([^"]+\.py)"', run_all.read_text(encoding="utf-8")):
            reference = Path(*relative).as_posix()
            if not (dataset_root / "raw" / reference).is_file():
                missing_script_references.append(reference)

    duplicates = {key: count for key, count in sorted(identifiers.items()) if count > 1}
    coordinate_fields = ["Y", "X", "Z"]
    coordinate_ranges = {field: _range(numeric_ranges[field]) for field in coordinate_fields}
    return {
        "profile": "coal_boreholes_602_v001",
        "content_classification": "structured_directional_coal_borehole_records",
        "source_data_role": "source_structured_data",
        "workbook": {
            "filename": filename,
            "sha256": inventory[filename]["sha256"],
            "sheet_name": sheet.title,
            "worksheet_dimensions": {"rows_including_headers": sheet.max_row, "columns": sheet.max_column},
            "merged_range_count": len(sheet.merged_cells.ranges),
            "formula_cell_count": formula_cells,
            "flattened_headers": headers,
            "headers_match_profile": headers == expected_headers,
        },
        "record_count": len(rows),
        "unique_borehole_id_count": len(identifiers),
        "duplicate_borehole_ids": duplicates,
        "missing_cells_by_field": dict(sorted(missing.items())),
        "numeric_ranges": {field: _range(values) for field, values in numeric_ranges.items()},
        "source_field_observations": {
            "roof_depth_greater_than_final_depth_count": roof_greater_than_final,
            "roof_depth_plus_seam_thickness_greater_than_final_depth_count": roof_plus_thickness_greater_than_final,
            "final_depth_minus_roof_depth_range_m": _range(final_minus_roof_values),
            "nonpositive_final_roof_or_thickness_count": nonpositive_depth_or_thickness,
            "maximum_abs_zenith_minus_inclination_minus_90_degrees": angle_relation_max_error,
            "maximum_abs_drilling_minus_borehole_azimuth_degrees": azimuth_relation_max_error,
            "interpretation_note": (
                "The roof-plus-thickness comparison is descriptive, not a validity constraint: "
                "the source contains directional drilled depths and coal-seam geometric fields "
                "whose reference definitions are not fully specified in the released files."
            ),
        },
        "spatial_metadata": {
            "numeric_coordinate_or_elevation_fields": coordinate_fields,
            "coordinate_ranges": coordinate_ranges,
            "coordinate_reference_system_declared_in_released_files": False,
            "precise_coordinate_values_present": True,
            "privacy_and_sensitive_location_review": "pending_human_review",
        },
        "released_code_audit": {
            "entrypoint_present": run_all.is_file(),
            "missing_entrypoint_script_references": sorted(missing_script_references),
            "entrypoint_runnable_as_released": not missing_script_references,
        },
        "eligibility": {
            "phase1_document_extraction": False,
            "paper1_benchmark": False,
            "paper2_method_evaluation": False,
            "paper3_protocol_development": True,
            "paper3_formal_spatial_error_propagation": False,
        },
        "formal_use_status": "candidate_pending_human_and_spatial_review",
        "formal_use_blockers": [
            "coordinate_reference_system_not_declared",
            "privacy_and_sensitive_location_review_pending",
            "source_is_not_ai_extraction_output_or_human_ground_truth",
            "released_entrypoint_references_missing_scripts",
            "paper3_experiment_protocol_not_preregistered_or_run",
        ],
    }


def build_structured_dataset_audit(
    dataset_root: Path,
    destination: Path,
    *,
    profile: str,
    audited_at_utc: str | None = None,
) -> dict[str, Any]:
    """Create an immutable audit directory bound to acquisition and raw-file hashes."""

    if profile not in SUPPORTED_PROFILES:
        raise ValueError(f"unsupported structured-audit profile: {profile}")
    if destination.exists():
        raise FileExistsError(f"refusing to overwrite structured audit: {destination}")
    acquisition, inventory = _load_acquisition(dataset_root)
    audit_body = (
        _audit_binhai(dataset_root, inventory)
        if profile == "binhai_cptu_v001"
        else _audit_coal(dataset_root, inventory)
    )
    privacy_signals = _scan_text_privacy_signals(dataset_root, inventory)
    payload = {
        "audit_schema_version": AUDIT_SCHEMA_VERSION,
        "audited_at_utc": audited_at_utc or datetime.now(timezone.utc).isoformat(),
        "dataset_id": acquisition["dataset_id"],
        "dataset_doi": acquisition["dataset_doi"],
        "dataset_version": acquisition["dataset_version"],
        "license_id": acquisition["license_id"],
        "acquisition_sha256": _sha256_file(dataset_root / "metadata/acquisition.json"),
        "source_files": [
            {
                "filename": name,
                "size_bytes": row["size_bytes"],
                "sha256": row["sha256"],
            }
            for name, row in sorted(inventory.items())
        ],
        "automated_privacy_screen": {
            "scope": "released text files and structured field presence; not a human privacy determination",
            "text_findings": privacy_signals,
            "human_review_completed": False,
        },
        "content_audit": audit_body,
    }

    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary = Path(tempfile.mkdtemp(prefix=f".{destination.name}.", dir=destination.parent))
    try:
        audit_path = temporary / "structured_content_audit.json"
        audit_path.write_bytes(_json_bytes(payload))
        manifest = {
            "manifest_schema_version": "structured_audit_artifact_manifest_v001",
            "files": [{
                "path": audit_path.name,
                "size_bytes": audit_path.stat().st_size,
                "sha256": _sha256_file(audit_path),
            }],
        }
        manifest_path = temporary / "artifact_manifest.json"
        manifest_path.write_bytes(_json_bytes(manifest))
        os.replace(temporary, destination)
    except Exception:
        shutil.rmtree(temporary, ignore_errors=True)
        raise
    return verify_structured_dataset_audit(dataset_root, destination)


def verify_structured_dataset_audit(dataset_root: Path, audit_root: Path) -> dict[str, Any]:
    """Verify audit artifacts and every bound acquired source file."""

    manifest_path = audit_root / "artifact_manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    listed = manifest.get("files")
    if not isinstance(listed, list) or not listed:
        raise ValueError("structured-audit artifact manifest is empty")
    expected_files = {"artifact_manifest.json"}
    for row in listed:
        relative = str(row.get("path") or "")
        if not relative or Path(relative).name != relative:
            raise ValueError("structured-audit manifest contains an unsafe path")
        path = audit_root / relative
        expected_files.add(relative)
        if path.stat().st_size != row.get("size_bytes") or _sha256_file(path) != row.get("sha256"):
            raise ValueError(f"structured-audit artifact mismatch: {relative}")
    actual_files = {path.name for path in audit_root.iterdir() if path.is_file()}
    if actual_files != expected_files:
        raise ValueError("structured-audit file set does not match artifact manifest")

    payload = json.loads((audit_root / "structured_content_audit.json").read_text(encoding="utf-8"))
    if payload.get("audit_schema_version") != AUDIT_SCHEMA_VERSION:
        raise ValueError("unsupported structured-audit schema version")
    acquisition, inventory = _load_acquisition(dataset_root)
    if payload.get("acquisition_sha256") != _sha256_file(dataset_root / "metadata/acquisition.json"):
        raise ValueError("structured audit is bound to a different acquisition")
    if payload.get("dataset_id") != acquisition.get("dataset_id"):
        raise ValueError("structured audit dataset ID mismatch")
    expected_sources = {
        (name, row["size_bytes"], row["sha256"])
        for name, row in inventory.items()
    }
    observed_sources = {
        (row.get("filename"), row.get("size_bytes"), row.get("sha256"))
        for row in payload.get("source_files", [])
    }
    if observed_sources != expected_sources:
        raise ValueError("structured audit source-file binding mismatch")
    return {
        "dataset_id": payload["dataset_id"],
        "profile": payload["content_audit"]["profile"],
        "formal_use_status": payload["content_audit"]["formal_use_status"],
        "source_file_count": len(expected_sources),
        "audit_sha256": _sha256_file(audit_root / "structured_content_audit.json"),
        "manifest_sha256": _sha256_file(manifest_path),
        "verified": True,
    }
