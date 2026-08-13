"""Protocol-only source-field surfaces from licensed structured workbooks."""

from __future__ import annotations

from dataclasses import dataclass
import math
from pathlib import Path
import random
from typing import Any, Iterable, Sequence

from .error_propagation import SurfacePoint


@dataclass(frozen=True)
class SourceFieldSurface:
    """Local-coordinate scalar points plus non-identifying audit summaries."""

    points: tuple[SurfacePoint, ...]
    source_record_count: int
    coordinate_extent_u_m: float
    coordinate_extent_v_m: float
    scalar_minimum_m: float
    scalar_maximum_m: float
    coordinate_origin_persisted: bool = False


COAL_602_HEADERS = (
    "Gas Drainage Borehole Number", "Y", "X", "Z", "Final hole depth(m)",
    "3 Coal Roof Depth(m)", "3 Coal seam thickness(m)", "Borehole Zenith Angle (°)",
    "Drilling azimuth (°)", "borehole inclina-tion (°)", "borehole azimuth (°)",
    "coal seam dip angle (°)", "coal seam dip direction (°)",
)


def _openpyxl():
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - environment-dependent guard
        raise RuntimeError("source-field workbook loading requires openpyxl") from exc
    return load_workbook


def _flatten_headers(sheet: Any) -> tuple[str, ...]:
    first = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    second = list(next(sheet.iter_rows(min_row=2, max_row=2, values_only=True)))
    return tuple(
        str((second[index] if index < len(second) and second[index] is not None
             else first[index] if index < len(first) else "") or "").strip()
        for index in range(max(len(first), len(second)))
    )


def _number(value: Any, *, field: str, row_number: int) -> float:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError(f"non-numeric {field} at workbook row {row_number}")
    result = float(value)
    if not math.isfinite(result):
        raise ValueError(f"non-finite {field} at workbook row {row_number}")
    return result


def load_coal_602_roof_depth_surface(workbook_path: Path) -> SourceFieldSurface:
    """Load roof-depth values on origin-suppressed source collar coordinates.

    The returned scalar is the source-reported roof depth. It is deliberately
    not converted to elevation because the directional-hole reference semantics
    and CRS are not established by the released files.
    """

    load_workbook = _openpyxl()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if len(workbook.sheetnames) != 1:
        raise ValueError("coal-602 source-field profile requires one worksheet")
    sheet = workbook[workbook.sheetnames[0]]
    headers = _flatten_headers(sheet)
    if headers != COAL_602_HEADERS:
        raise ValueError("coal-602 workbook header does not match frozen profile")
    source_rows: list[tuple[str, float, float, float]] = []
    seen_ids: set[str] = set()
    for row_number, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), 3):
        borehole_id = str(row[0] or "").strip()
        if not borehole_id or borehole_id in seen_ids:
            raise ValueError(f"missing or duplicate borehole ID at workbook row {row_number}")
        seen_ids.add(borehole_id)
        # Preserve the source labels: local u is source Y and local v is source X.
        source_u = _number(row[1], field="source Y", row_number=row_number)
        source_v = _number(row[2], field="source X", row_number=row_number)
        roof_depth = _number(row[5], field="roof depth", row_number=row_number)
        source_rows.append((borehole_id, source_u, source_v, roof_depth))
    workbook.close()
    if len(source_rows) < 3:
        raise ValueError("source-field surface requires at least three records")
    minimum_u = min(row[1] for row in source_rows)
    minimum_v = min(row[2] for row in source_rows)
    maximum_u = max(row[1] for row in source_rows)
    maximum_v = max(row[2] for row in source_rows)
    scalar_values = [row[3] for row in source_rows]
    points = tuple(
        SurfacePoint(
            x=source_u - minimum_u,
            y=source_v - minimum_v,
            elevation=roof_depth,
            borehole_id=f"SOURCE_ROW_{index:04d}",
        )
        for index, (_, source_u, source_v, roof_depth) in enumerate(source_rows, 1)
    )
    if len({(point.x, point.y) for point in points}) != len(points):
        raise ValueError("source-field profile contains duplicate collar coordinates")
    return SourceFieldSurface(
        points=points,
        source_record_count=len(points),
        coordinate_extent_u_m=maximum_u - minimum_u,
        coordinate_extent_v_m=maximum_v - minimum_v,
        scalar_minimum_m=min(scalar_values),
        scalar_maximum_m=max(scalar_values),
    )


def convex_hull_xy(points: Iterable[SurfacePoint]) -> tuple[tuple[float, float], ...]:
    """Return the counter-clockwise monotonic-chain hull without repetition."""

    coordinates = sorted({(float(point.x), float(point.y)) for point in points})
    if len(coordinates) < 3:
        raise ValueError("convex hull requires at least three unique points")

    def cross(origin: tuple[float, float], a: tuple[float, float], b: tuple[float, float]) -> float:
        return (a[0] - origin[0]) * (b[1] - origin[1]) - (a[1] - origin[1]) * (b[0] - origin[0])

    lower: list[tuple[float, float]] = []
    for point in coordinates:
        while len(lower) >= 2 and cross(lower[-2], lower[-1], point) <= 0:
            lower.pop()
        lower.append(point)
    upper: list[tuple[float, float]] = []
    for point in reversed(coordinates):
        while len(upper) >= 2 and cross(upper[-2], upper[-1], point) <= 0:
            upper.pop()
        upper.append(point)
    hull = tuple(lower[:-1] + upper[:-1])
    if len(hull) < 3:
        raise ValueError("source coordinates are collinear")
    return hull


def _inside_convex_hull(
    point: tuple[float, float], hull: Sequence[tuple[float, float]], tolerance: float = 1e-9,
) -> bool:
    signs = []
    for index, start in enumerate(hull):
        end = hull[(index + 1) % len(hull)]
        cross = (end[0] - start[0]) * (point[1] - start[1]) - (end[1] - start[1]) * (point[0] - start[0])
        if abs(cross) > tolerance:
            signs.append(cross > 0)
    return not signs or all(sign == signs[0] for sign in signs)


def regular_queries_within_hull(
    hull: Sequence[tuple[float, float]], grid_size: int,
) -> tuple[tuple[float, float], ...]:
    """Build a deterministic regular grid and retain convex-hull interior cells."""

    if grid_size < 3:
        raise ValueError("grid_size must be at least 3")
    if len(hull) < 3:
        raise ValueError("query hull requires at least three vertices")
    minimum_x, maximum_x = min(x for x, _ in hull), max(x for x, _ in hull)
    minimum_y, maximum_y = min(y for _, y in hull), max(y for _, y in hull)
    if maximum_x == minimum_x or maximum_y == minimum_y:
        raise ValueError("query hull must have non-zero extent")
    x_values = [minimum_x + index * (maximum_x - minimum_x) / (grid_size - 1) for index in range(grid_size)]
    y_values = [minimum_y + index * (maximum_y - minimum_y) / (grid_size - 1) for index in range(grid_size)]
    queries = tuple(
        (x, y) for x in x_values for y in y_values
        if _inside_convex_hull((x, y), hull)
    )
    if not queries:
        raise ValueError("query grid contains no convex-hull points")
    return queries


def perturb_surface_scalar(
    points: Sequence[SurfacePoint], magnitude_m: float, seed: int,
) -> tuple[SurfacePoint, ...]:
    """Apply independent fixed-magnitude signed perturbations to scalar values."""

    if magnitude_m < 0:
        raise ValueError("perturbation magnitude must be non-negative")
    rng = random.Random(seed)
    return tuple(
        SurfacePoint(
            x=point.x,
            y=point.y,
            elevation=point.elevation + (rng.choice((-magnitude_m, magnitude_m)) if magnitude_m else 0.0),
            borehole_id=point.borehole_id,
        )
        for point in points
    )
